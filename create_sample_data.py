"""
create_sample_data.py
生成一份示例客户 Excel 表格，用于测试"AI 门店军师"。
运行方式：python create_sample_data.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_sample_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "客户信息"

    # ---------- 表头样式 ----------
    header_font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["客户姓名", "车型", "上次到店时间", "上次维修项目"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ---------- 示例数据 ----------
    # 包含不同时间跨度的客户，方便测试"超半年"筛选逻辑
    sample_data = [
        # ---- 超过半年未到店（应被筛选出来）----
        ["张三",   "宝马 3系 2020款",   "2025-03-10", "更换机油+机滤"],
        ["李四",   "丰田凯美瑞 2021款", "2025-05-22", "四轮定位"],
        ["王五",   "奔驰 C200L",       "2025-01-15", "更换刹车片"],
        ["赵六",   "大众帕萨特 2019款", "2025-06-01", "空调清洗+制冷剂补充"],
        ["孙七",   "本田雅阁 2022款",   "2024-12-08", "全车漆面镀晶"],

        # ---- 半年内到过店（不应被筛选出来）----
        ["周八",   "奥迪 A4L 2023款",   "2025-11-20", "常规保养"],
        ["吴九",   "特斯拉 Model 3",    "2026-01-05", "底盘检查"],
        ["郑十",   "比亚迪汉 EV",       "2026-02-28", "轮胎更换"],
    ]

    data_font = Font(name="微软雅黑", size=11)
    data_alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    # ---------- 自动调整列宽 ----------
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 26

    # ---------- 保存 ----------
    filepath = "客户信息表.xlsx"
    wb.save(filepath)
    print(f"示例 Excel 已生成：{filepath}")
    print(f"共 {len(sample_data)} 条客户记录")
    print(f"其中约 5 条超过半年未到店，3 条为近期客户")

if __name__ == "__main__":
    create_sample_excel()
