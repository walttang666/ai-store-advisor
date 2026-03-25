"""
AI 门店军师 MVP - 主程序
========================================
功能：读取客户 Excel → 筛选超半年未到店 → AI 生成促活话术 → 输出结果 Excel

运行方式：python main.py
首次运行前请先：
  1. 运行 python create_sample_data.py 生成示例数据（或放入你自己的 Excel）
  2. 编辑 .env 文件，填入你的 AI API 密钥
"""

import os
import sys
import json
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from dotenv import load_dotenv

# 脚本所在目录，确保无论从哪个目录运行都能找到数据文件
_BASE_DIR = Path(__file__).parent

import database

# ============================================================
# 第 1 阶段：读取 Excel
# ============================================================

def read_customer_excel(filepath):
    """
    读取客户 Excel 文件，返回客户列表。
    每个客户是一个字典：{姓名, 车型, 上次到店时间, 上次维修项目}
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"找不到文件 '{filepath}'")

    wb = load_workbook(filepath)
    ws = wb.active

    # 读取表头（第一行），找到各列的位置
    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[cell.value.strip()] = col_idx

    # 检查必需的列是否存在（通过别名机制）
    alias_map = {
        "客户姓名": ["客户姓名", "姓名", "车主姓名", "客户名"],
        "车型": ["车型", "车辆型号", "车系", "车牌及车型"],
        "上次到店时间": ["上次到店时间", "上次进站时间", "最近一次到店时间", "进站时间", "最后到店时间"],
        "上次维修项目": ["上次维修项目", "维修项目", "服务项目", "上次服务内容", "上次消费项目"]
    }
    
    # 找到匹配的实际列名
    matched_headers = {}
    missing = []
    
    for key, aliases in alias_map.items():
        found = False
        for alias in aliases:
            if alias in headers:
                matched_headers[key] = headers[alias]
                found = True
                break
        if not found:
            missing.append(key)
            
    if missing:
        raise ValueError(f"Excel 缺少关于 [{', '.join(missing)}] 的列名。您的表包含: {list(headers.keys())}")

    # 逐行读取数据
    customers = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        name = ws.cell(row=row[0].row, column=matched_headers["客户姓名"]).value
        car = ws.cell(row=row[0].row, column=matched_headers["车型"]).value
        last_visit = ws.cell(row=row[0].row, column=matched_headers["上次到店时间"]).value
        last_service = ws.cell(row=row[0].row, column=matched_headers["上次维修项目"]).value

        if not name:  # 跳过空行
            continue

        # 处理日期：支持多种常见中文 Excel 时间格式
        if isinstance(last_visit, str):
            last_visit_str = last_visit.strip()
            last_visit_str = last_visit_str.replace('/', '-').replace('.', '-').split(' ')[0] # 统一成分隔符只留日期
            try:
                last_visit = datetime.strptime(last_visit_str, "%Y-%m-%d")
            except ValueError:
                try:
                    # 尝试 2024-1-1 这种非零补位或是其他奇怪的
                    last_visit = datetime.strptime(last_visit_str, "%Y-%m-%d")
                except ValueError:
                    print(f"警告：客户 '{name}' 的时间格式过奇葩: {last_visit}，已跳过")
                    continue
        elif isinstance(last_visit, datetime):
            pass  # 已经是 datetime，无需转换
        else:
            print(f"警告：客户 '{name}' 的到店时间为空或特殊对象，已跳过")
            continue

        customers.append({
            "姓名": name,
            "车型": car,
            "上次到店时间": last_visit,
            "上次维修项目": last_service,
        })

    wb.close()
    
    # 存入本地数据库
    database.init_db()
    for c in customers:
        database.upsert_customer(
            c["姓名"], 
            c["车型"], 
            c["上次到店时间"].strftime("%Y-%m-%d"), 
            c["上次维修项目"]
        )
        
    print(f"   已将 {len(customers)} 条记录同步至本地数据库。")
    return customers


# ============================================================
# 第 2 阶段：筛选超过半年未到店的客户
# ============================================================

def filter_inactive_customers(customers, days_threshold=180):
    """
    筛选出超过指定天数未到店的客户。
    默认阈值：180 天（约半年）。
    """
    today = datetime.now()
    inactive = []

    for c in customers:
        days_since = (today - c["上次到店时间"]).days
        if days_since >= days_threshold:
            c["未到店天数"] = days_since
            inactive.append(c)

    # 按未到店天数从大到小排序（最久没来的排最前）
    inactive.sort(key=lambda x: x["未到店天数"], reverse=True)
    return inactive


# ============================================================
# 第 3 阶段：调用 AI 生成促活话术
# ============================================================

def load_store_info():
    """读取本地门店配置"""
    try:
        with open(_BASE_DIR / "store_info.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"警告：无法读取门店配置 ({e})，使用默认空白信息。")
        return {"store_name": "本服务站", "store_address": "", "contact_phone": "", "current_promotion": ""}

def generate_reactivation_message(customer, client, store_info):
    """
    调用 AI 大模型，为单个客户生成一条定制化促活话术。
    """
    model = os.getenv("AI_MODEL", "deepseek-chat")

    prompt = f"""你是一位在一线干了十年的老汽修师/接待员，请给下面的车主发一条微信，叫他有空过来趟店里。
你的店名：{store_info.get('store_name')}
你们近期的活动：{store_info.get('current_promotion')}

车主信息：
- 称呼：{customer['姓名']}（男士叫哥，女士叫姐，不确定就直接叫名字，如果你觉得判断不准就叫老板，**绝对不要**用“先生/女士”和“您好”）
- 车：{customer['车型']}
- 上次来是 {customer['上次到店时间'].strftime('%Y年%m月%d日')}，做的项目是：{customer['上次维修项目']}
- 算下来已经有 {customer['未到店天数']} 天没来了

你的说话要求：
1. 极其接地气，像熟人唠嗑，一上来就直接打招呼。比如“王哥，最近忙啥呢”。
2. **严禁**使用套话，比如“咱们”、“专属售后顾问”、“期待您的光临”这类词汇，太假了。
3. 提醒他差不多该保养或者检查下车了，说出车况隐患（结合他上次的项目和车型）。
4. 顺势提一嘴你们店最近的活动（参考上面信息栏，别生硬，像顺口一提，比如“刚好最近咱们店里搞...你有空顺道过来薅个羊毛”）。也可以留个你店地址：{store_info.get('store_address')}。
5. 字数控制在100字左右，短平快！
6. 只输出正文内容。"""

    try:
        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": "你是个实在、热心肠的老汽修人。不讲官话，像跟左邻右舍打招呼一样给老客户发微信。"},
                {"role": "user", "content": prompt}
            ],
            temperature=0.7,
            max_tokens=200,
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        return f"[话术生成失败：{str(e)}]"

def ai_helper_generate_promotion(ideas_text, store_info, client):
    """
    AI 营销企划：根据用户口语化的草稿想法，结合店名，生成 3 套专业的促销方案供老板挑选。
    """
    model = os.getenv("AI_MODEL", "deepseek-chat")
    
    prompt = f"""你是一名顶级的“汽车后市场营销总监”。你的店名是【{store_info.get('store_name')}】。
现在，店面老板给你发了一段他粗糙的促销想法：
“{ideas_text}”

请帮他整理并扩写成 3 套结构清晰、文字吸引眼球的促销方案（侧重：简单粗暴、利他性强、引流效果好）。
每套方案需要包含：
1. 吸引人的霸气标题
2. 具体的优惠明细（例如原价多少、现价多少，送什么）
3. 方案的营销亮点说明（为什么能吸引客户）

请直接输出 3 套方案，不需要任何前言和总结。"""

    try:
        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": "你是一位顶级汽修店营销总监，擅长策划引流拓客活动。"},
                {"role": "user", "content": prompt}
            ],
            temperature=0.8,
            max_tokens=800,
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        return f"[企划生成失败：{str(e)}]"

def generate_all_messages(inactive_customers):
    """
    为所有筛选出的客户批量生成话术。
    """
    # 加载 .env 配置（使用绝对路径，兼容任意工作目录，并强制覆盖以防 Streamlit 缓存）
    load_dotenv(dotenv_path=_BASE_DIR / ".env", override=True)

    api_key = os.getenv("AI_API_KEY")
    base_url = os.getenv("AI_BASE_URL")

    if not api_key or api_key == "在这里粘贴你的API密钥":
        print("错误：请先在 .env 文件中填入你的 AI API 密钥！")
        print("   打开项目文件夹下的 .env 文件，将 AI_API_KEY= 后面的内容替换为真实密钥")
        sys.exit(1)

    from openai import OpenAI
    client = OpenAI(api_key=api_key, base_url=base_url)

    store_info = load_store_info()
    print(f"\n正在调用 AI 生成话术（共 {len(inactive_customers)} 位客户，融合【{store_info.get('store_name')}】店面信息）...\n")

    for i, customer in enumerate(inactive_customers, 1):
        print(f"   [{i}/{len(inactive_customers)}] 正在为 {customer['姓名']} 生成话术...")
        message = generate_reactivation_message(customer, client, store_info)
        customer["促活话术"] = message
        print(f"   {customer['姓名']} - 完成")

    print(f"\n全部话术生成完毕！")
    return inactive_customers


# ============================================================
# V3 新增：后台支持中断与异步的多线程生成模型
# ============================================================
import threading

bg_generation_status = {
    "is_running": False,
    "total": 0,
    "current": 0,
    "stop_requested": False,
    "results": [],
    "logs": []
}

def bg_generate_worker(inactive_customers, client, store_info):
    global bg_generation_status
    bg_generation_status["is_running"] = True
    bg_generation_status["stop_requested"] = False
    bg_generation_status["total"] = len(inactive_customers)
    bg_generation_status["current"] = 0
    bg_generation_status["results"] = []
    bg_generation_status["logs"] = []
    
    for c in inactive_customers:
        if bg_generation_status["stop_requested"]:
            bg_generation_status["logs"].append("⚠️ 已收到停止指令，生成终止。")
            break
            
        msg = generate_reactivation_message(c, client, store_info)
        c['促活话术'] = msg
        bg_generation_status["results"].append(c)
        bg_generation_status["current"] += 1
        bg_generation_status["logs"].append(f"✅ {c['姓名']} 话术已生成。")
        
        # V3 Sprint 3: 落盘到数据库历史记录表
        database.save_message_draft(c['姓名'], msg)
        
    bg_generation_status["is_running"] = False


# ============================================================
# 第 4 阶段：输出结果 Excel
# ============================================================

def save_results_to_excel(results, output_path="促活话术结果.xlsx"):
    """
    将生成的话术结果保存为新的 Excel 文件。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "促活话术"

    # ---------- 表头 ----------
    headers = ["客户姓名", "车型", "上次到店时间", "上次维修项目", "未到店天数", "AI 生成促活话术"]
    header_font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ---------- 数据行 ----------
    data_font = Font(name="微软雅黑", size=11)
    wrap_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, customer in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=customer["姓名"]).font = data_font
        ws.cell(row=row_idx, column=1).alignment = center_alignment
        ws.cell(row=row_idx, column=1).border = thin_border

        ws.cell(row=row_idx, column=2, value=customer["车型"]).font = data_font
        ws.cell(row=row_idx, column=2).alignment = center_alignment
        ws.cell(row=row_idx, column=2).border = thin_border

        ws.cell(row=row_idx, column=3, value=customer["上次到店时间"].strftime("%Y-%m-%d")).font = data_font
        ws.cell(row=row_idx, column=3).alignment = center_alignment
        ws.cell(row=row_idx, column=3).border = thin_border

        ws.cell(row=row_idx, column=4, value=customer["上次维修项目"]).font = data_font
        ws.cell(row=row_idx, column=4).alignment = center_alignment
        ws.cell(row=row_idx, column=4).border = thin_border

        ws.cell(row=row_idx, column=5, value=customer["未到店天数"]).font = data_font
        ws.cell(row=row_idx, column=5).alignment = center_alignment
        ws.cell(row=row_idx, column=5).border = thin_border

        ws.cell(row=row_idx, column=6, value=customer["促活话术"]).font = data_font
        ws.cell(row=row_idx, column=6).alignment = wrap_alignment
        ws.cell(row=row_idx, column=6).border = thin_border

    # ---------- 调整列宽 ----------
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 55

    # 自动调整话术行高
    for row_idx in range(2, len(results) + 2):
        ws.row_dimensions[row_idx].height = 80

    wb.save(output_path)
    return output_path


# ============================================================
# 主流程
# ============================================================

def main():
    print("=" * 55)
    print("   AI 门店军师 MVP - 客户促活话术生成器")
    print("=" * 55)

    # ---------- 第 1 步：读取 Excel ----------
    input_file = _BASE_DIR / "客户信息表.xlsx"
    print(f"\n第 1 步：读取客户数据 ({input_file})...")
    customers = read_customer_excel(input_file)
    print(f"   共读取到 {len(customers)} 条客户记录")

    # ---------- 第 2 步：筛选 ----------
    print(f"\n第 2 步：筛选超过半年未到店的客户...")
    inactive = filter_inactive_customers(customers)
    print(f"   共发现 {len(inactive)} 位客户超过半年未到店：")
    for c in inactive:
        print(f"      - {c['姓名']}（{c['车型']}）- 已 {c['未到店天数']} 天未到店")

    if not inactive:
        print("\n太好了！所有客户都在半年内到过店，无需生成促活话术。")
        return

    # ---------- 第 3 步：生成话术 ----------
    print(f"\n第 3 步：调用 AI 生成促活话术...")
    results = generate_all_messages(inactive)

    # ---------- 第 4 步：输出结果 ----------
    print(f"\n第 4 步：保存结果到 Excel...")
    output_file = save_results_to_excel(results, str(_BASE_DIR / "促活话术结果.xlsx"))
    print(f"   结果已保存到：{output_file}")

    # ---------- 运行摘要 ----------
    print(f"\n{'=' * 55}")
    print(f"   运行摘要")
    print(f"{'=' * 55}")
    print(f"   总客户数：  {len(customers)} 人")
    print(f"   需促活数：  {len(inactive)} 人")
    print(f"   话术已生成：{len(results)} 条")
    print(f"   输出文件：  {output_file}")
    print(f"{'=' * 55}")
    print(f"\n   请打开 '{output_file}' 查看话术，")
    print(f"   逐条复制发送给对应客户即可！")


if __name__ == "__main__":
    main()
